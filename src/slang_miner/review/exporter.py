"""人工审核 xlsx 导出 / 导入模块。

负责把流水线产出的 :class:`SlangEntry` 列表导出成「周报」风格的 Excel
（带表头样式、冻结首行、人工判定下拉数据验证），供人工逐条审核；并能把
人工填好的 xlsx 读回成 :class:`SlangEntry` 列表，解析其中的人工判定状态，
为下游闭环回灌（feedback）提供输入。

设计要点：
- 仅依赖 openpyxl，Python 3.9 兼容（统一使用 ``typing.List`` 等显式泛型）。
- 列顺序固定，导出与导入共用同一份列定义常量，避免「导出列改了导入读错位」。
- 导入侧对人工填写做容错与归一化：大小写、中英文别名、留空 → pending。
- 全程不修改入参对象，导入时生成新的 SlangEntry（遵循不可变风格）。
"""

from __future__ import annotations

from pathlib import Path
from typing import Dict, List, Union

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from slang_miner.schema import SlangEntry

# ---------------------------------------------------------------------------
# 列定义（导出 / 导入唯一真相源）
# ---------------------------------------------------------------------------

# 人工判定合法取值（写入下拉、导入校验都依赖它）。
# 留空表示尚未审核（pending），不在下拉选项里但允许存在。
VALID_VERDICTS = ("correct", "modified", "incorrect")

# 合法 status 全集（含 pending），用于导入侧最终校验。
VALID_STATUSES = ("pending", "correct", "modified", "incorrect")

# 表头文案 → 含义。顺序即为 Excel 中的列顺序，1-based 列号由列表下标推导。
HEADERS: List[str] = [
    "词term",
    "类别",
    "释义",
    "原文例句",
    "置信度",
    "来源",
    "人工判定(correct/modified/incorrect)",
    "修改后释义",
    "备注",
]

# 各列默认宽度（字符数），与 HEADERS 一一对应。
COLUMN_WIDTHS: List[int] = [16, 12, 40, 40, 10, 18, 28, 40, 24]

# 多来源在单元格内的分隔符（导出 join / 导入 split 共用）。
SOURCE_SEP = ","

# 人工判定别名归一化表：把常见中文 / 大小写写法映射到标准英文取值。
_VERDICT_ALIASES: Dict[str, str] = {
    "correct": "correct",
    "正确": "correct",
    "对": "correct",
    "modified": "modified",
    "修改": "modified",
    "已修改": "modified",
    "incorrect": "incorrect",
    "错误": "incorrect",
    "错": "incorrect",
    "不对": "incorrect",
}


# ---------------------------------------------------------------------------
# 样式常量
# ---------------------------------------------------------------------------

_HEADER_FILL = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
_HEADER_FONT = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
_CELL_ALIGN = Alignment(horizontal="left", vertical="top", wrap_text=True)
_THIN = Side(style="thin", color="D9D9D9")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


# ---------------------------------------------------------------------------
# 导出
# ---------------------------------------------------------------------------


def export_review_xlsx(entries: List[SlangEntry], path: Union[str, Path]) -> str:
    """把待审核词条导出为「周报」风格 xlsx。

    生成的工作表包含 9 列（见 :data:`HEADERS`），表头加底色 / 加粗 / 居中样式，
    冻结首行，并在「人工判定」列加下拉数据验证（correct/modified/incorrect，
    允许留空）。

    Args:
        entries: 待导出的黑话词条列表，通常 status 为 pending。
        path: 输出 xlsx 的绝对路径（建议绝对路径，父目录会自动创建）。

    Returns:
        实际写入的文件绝对路径字符串。

    Raises:
        TypeError: entries 中存在非 SlangEntry 元素时。
    """
    out_path = Path(path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "待审核黑话"

    _write_header(sheet)
    _write_rows(sheet, entries)
    _apply_layout(sheet, row_count=len(entries))
    _attach_verdict_validation(sheet, row_count=len(entries))

    workbook.save(str(out_path))
    return str(out_path.resolve())


def _write_header(sheet) -> None:
    """写入表头并套用样式。"""
    for col_idx, title in enumerate(HEADERS, start=1):
        cell = sheet.cell(row=1, column=col_idx, value=title)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _HEADER_ALIGN
        cell.border = _BORDER


def _write_rows(sheet, entries: List[SlangEntry]) -> None:
    """逐条写入数据行。

    列顺序严格对应 :data:`HEADERS`。
    「人工判定 / 修改后释义 / 备注」三列默认留空，等待人工填写；但若传入的
    entry 已带非 pending 状态（如二次导出复核），则把 status 回填到判定列。
    """
    for offset, entry in enumerate(entries):
        if not isinstance(entry, SlangEntry):
            raise TypeError(
                f"entries[{offset}] 期望 SlangEntry，实际为 {type(entry).__name__}"
            )

        row = offset + 2  # 第 1 行是表头
        sources_text = SOURCE_SEP.join(entry.sources or [])
        # 已有结论才回填判定列，pending 留空便于人工识别「未处理」。
        verdict_text = entry.status if entry.status in VALID_VERDICTS else ""

        values = [
            entry.term,
            entry.category,
            entry.definition,
            entry.example,
            entry.confidence,
            sources_text,
            verdict_text,
            "",  # 修改后释义：人工填
            "",  # 备注：人工填
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = sheet.cell(row=row, column=col_idx, value=value)
            cell.alignment = _CELL_ALIGN
            cell.border = _BORDER


def _apply_layout(sheet, row_count: int) -> None:
    """设置列宽、冻结首行。"""
    for col_idx, width in enumerate(COLUMN_WIDTHS, start=1):
        sheet.column_dimensions[get_column_letter(col_idx)].width = width
    # 冻结首行：A2 表示「第 2 行及以后可滚动，第 1 行固定」。
    sheet.freeze_panes = "A2"


def _attach_verdict_validation(sheet, row_count: int) -> None:
    """在「人工判定」列挂下拉数据验证（correct/modified/incorrect）。

    allow_blank=True 允许留空（=尚未审核）。即便当前没有数据行，也对一段
    预留区间挂上验证，方便人工手动追加行时仍有下拉。
    """
    verdict_col = HEADERS.index("人工判定(correct/modified/incorrect)") + 1
    col_letter = get_column_letter(verdict_col)

    formula = '"{}"'.format(",".join(VALID_VERDICTS))
    validation = DataValidation(
        type="list",
        formula1=formula,
        allow_blank=True,
        showDropDown=False,  # False = 显示下拉箭头（openpyxl 语义反直觉，勿改）
    )
    validation.error = "请从下拉中选择 correct / modified / incorrect，或留空表示未审核"
    validation.errorTitle = "无效的人工判定"
    validation.prompt = "correct=正确; modified=释义需修改(请填修改后释义); incorrect=不是黑话"
    validation.promptTitle = "人工判定"

    # 覆盖现有数据行 + 预留若干空行，便于人工追加。
    last_row = max(row_count + 1, 1) + 50
    validation.add("{0}2:{0}{1}".format(col_letter, last_row))
    sheet.add_data_validation(validation)


# ---------------------------------------------------------------------------
# 导入
# ---------------------------------------------------------------------------


def import_verdicts(path: Union[str, Path]) -> List[SlangEntry]:
    """读回人工填好的 xlsx，解析为 :class:`SlangEntry` 列表。

    解析规则：
    - 按表头文案定位列，不依赖硬编码列号（人工若插列也能容错）。
    - 「人工判定」列经别名归一化后决定 status：
        * correct  → status="correct"
        * modified → status="modified"（若填了「修改后释义」，覆盖 definition）
        * incorrect→ status="incorrect"
        * 留空 / 无法识别 → status="pending"
    - 「来源」列按逗号拆分还原为 sources 列表。
    - 跳过完全空白行与缺少 term 的行。

    Args:
        path: 人工填好的 xlsx 绝对路径。

    Returns:
        解析得到的 SlangEntry 列表（每个都是新对象）。

    Raises:
        FileNotFoundError: 文件不存在时。
        ValueError: 找不到必需的表头列时。
    """
    in_path = Path(path)
    if not in_path.exists():
        raise FileNotFoundError(f"待导入的审核文件不存在：{in_path}")

    workbook = load_workbook(filename=str(in_path), read_only=True, data_only=True)
    try:
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        try:
            header_row = next(rows)
        except StopIteration:
            return []  # 空表

        col_index = _build_header_index(header_row)
        entries: List[SlangEntry] = []
        for raw in rows:
            entry = _parse_row(raw, col_index)
            if entry is not None:
                entries.append(entry)
        return entries
    finally:
        workbook.close()


def _build_header_index(header_row) -> Dict[str, int]:
    """把表头行映射为「表头文案 → 0-based 列下标」。

    必需列：词term、类别、释义、原文例句、置信度、来源、人工判定。
    「修改后释义」「备注」为可选列（缺失则按空处理）。
    """
    normalized = {
        (str(value).strip() if value is not None else ""): idx
        for idx, value in enumerate(header_row)
    }

    required = HEADERS[:7]  # 前 7 列为必需
    missing = [h for h in required if h not in normalized]
    if missing:
        raise ValueError(f"xlsx 缺少必需的表头列：{missing}")

    return normalized


def _parse_row(raw, col_index: Dict[str, int]) -> Union[SlangEntry, None]:
    """把一行原始单元格元组解析为 SlangEntry；空行 / 无 term 行返回 None。"""
    if raw is None or all(cell is None for cell in raw):
        return None

    term = _cell(raw, col_index, "词term")
    if not term:
        return None  # 无词条的行视为无效

    category = _cell(raw, col_index, "类别")
    definition = _cell(raw, col_index, "释义")
    example = _cell(raw, col_index, "原文例句")
    confidence = _to_float(_cell_raw(raw, col_index, "置信度"))
    sources = _split_sources(_cell(raw, col_index, "来源"))

    verdict_raw = _cell(raw, col_index, "人工判定(correct/modified/incorrect)")
    status = _normalize_verdict(verdict_raw)

    # modified 且人工填了「修改后释义」→ 用修改后的释义覆盖原释义。
    modified_def = _cell(raw, col_index, "修改后释义")
    if status == "modified" and modified_def:
        definition = modified_def

    return SlangEntry(
        term=term,
        category=category,
        definition=definition,
        example=example,
        confidence=confidence,
        sources=sources,
        status=status,
    )


def _cell_raw(raw, col_index: Dict[str, int], header: str):
    """取某列原始值（可能为 None / 数字 / 字符串）；列不存在返回 None。"""
    idx = col_index.get(header)
    if idx is None or idx >= len(raw):
        return None
    return raw[idx]


def _cell(raw, col_index: Dict[str, int], header: str) -> str:
    """取某列并转为去空白字符串；None → 空串。"""
    value = _cell_raw(raw, col_index, header)
    if value is None:
        return ""
    return str(value).strip()


def _to_float(value) -> float:
    """把置信度单元格转 float；非法 / 空 → 0.0。"""
    if value is None or value == "":
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def _split_sources(text: str) -> List[str]:
    """把来源字符串按逗号拆分并去空白；空串 → 空列表。"""
    if not text:
        return []
    return [part.strip() for part in text.split(SOURCE_SEP) if part.strip()]


def _normalize_verdict(text: str) -> str:
    """把人工判定文本归一化为合法 status；无法识别 / 留空 → pending。"""
    if not text:
        return "pending"
    key = text.strip().lower()
    # 别名表的 key 既有英文小写也有中文；中文不受 lower 影响。
    mapped = _VERDICT_ALIASES.get(key) or _VERDICT_ALIASES.get(text.strip())
    if mapped in VALID_VERDICTS:
        return mapped
    return "pending"
